import os 
import uuid
import cloudinary
import cloudinary.uploader
import cloudinary.api
from werkzeug.utils import secure_filename
from flask import Flask,render_template,request,redirect,session,flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_
from werkzeug.security import (generate_password_hash,check_password_hash)
from flask_migrate import Migrate
from openpyxl import Workbook
from io import BytesIO
from flask import send_file
from datetime import datetime, timezone
from openpyxl.styles import Font
from datetime import datetime
from dotenv import load_dotenv



load_dotenv()
app = Flask(__name__)
UPLOAD_FOLDER = 'static/uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.secret_key = os.getenv('SECRET_KEY')
app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('DATABASE_URL')
db = SQLAlchemy(app)
migrate = Migrate(app,db)
from flask_migrate import upgrade

with app.app_context():
    upgrade()
cloudinary.config(
    cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET")
)
if not app.secret_key:
    raise ValueError("SECRET_KEY environment variable is not set")


class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_no = db.Column(db.String(20), nullable = False,unique=True)
    first_name = db.Column(db.String(100), nullable = False)
    last_name = db.Column(db.String(100), nullable = False)
    email = db.Column(db.String(100), nullable = False)
    department = db.Column(db.String(100), nullable = False)
    position = db.Column(db.String(100), nullable = False)
    status = db.Column(db.String(20), nullable = False)
    photo = db.Column(db.String(255), nullable=True)
    photo_source = db.Column(db.String(20), nullable=True,default="local")
    cloudinary_public_id = db.Column(db.String(255), nullable =True)

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), nullable=False,unique=True)
    password =db.Column(db.String(255),nullable=False)

class AuditLog(db.Model):
    id = db.Column(
        db.Integer, 
        primary_key=True
    )
    username = db.Column(
        db.String(255),
        nullable=False
    )
    action = db.Column(
        db.String(255),
        nullable=False
    )
    created_at = db.Column(
        db.DateTime, 
        default=datetime.now(timezone.utc)  

    )
    
def add_audit_log(action):

    log = AuditLog(
        username=session['username'],
        action=action
    )

    db.session.add(log)

@app.route('/audit-logs')
def audit_logs():

    if 'username' not in session:
        return redirect('/login')
    
    page = request.args.get(
        'page',
        1,
        type=int
    )

    logs = AuditLog.query.order_by(
        AuditLog.id.desc()
    ).paginate(
        page=page,
        per_page=10
    )

    return render_template(
        'audit_logs.html',
        logs=logs
    )

@app.route('/')
def home():
    if 'username' not in session:
        return redirect('/login')
    
    total_employees = Employee.query.count()
    active_employees = Employee.query.filter_by(status='Active').count()
    inactive_employees = Employee.query.filter_by(status='Inactive').count()

    page = request.args.get(
        'page',
        1,
        type=int
    )


    search = request.args.get('search','')
    if search:
        employees = Employee.query.filter(
        or_(
            Employee.employee_no.contains(search),
            Employee.first_name.contains(search),
            Employee.last_name.contains(search),
            Employee.email.contains(search),
            Employee.department.contains(search),
            Employee.position.contains(search),
            Employee.status.contains(search)
            )
        ).paginate(page=page,per_page=5)
    else:
        #employees = Employee.query.all()
        employees = Employee.query.paginate(
            page=page,
            per_page=5
        )
        
    for employee in employees.items:
        print("==========================")
        print("Employee No:", employee.employee_no)
        print("Photo Source:", employee.photo_source)
        print("Photo:", employee.photo) 
        
    return render_template(
        'index.html',
        employees=employees,
        search=search,
        username=session['username'],
        total_employees = total_employees,
        active_employees = active_employees,
        inactive_employees = inactive_employees
    )

@app.route('/add', methods=['POST'])
def add_employee():

    if 'username' not in session:
        return redirect('/login')

    employee_no = request.form['employee_no']
    first_name = request.form['first_name']
    last_name = request.form['last_name']
    email = request.form['email']
    department = request.form['department']
    position = request.form['position']
    status = request.form['status']
    photo = request.files['photo']

    # ============================
    # Check duplicate employee no
    # ============================

    existing_employee = Employee.query.filter_by(
        employee_no=employee_no
    ).first()

    if existing_employee:

        flash(
            "Employee Number already exists!",
            "danger"
        )

        return redirect("/")

    # ============================
    # Upload Photo
    # ============================

    filename = None
    photo_source = None
    cloudinary_public_id = None

    is_cloud = os.getenv(
        "IS_CLOUD",
        ""
    ).strip().lower()

    if photo and photo.filename:

        if is_cloud == "true":

            result = cloudinary.uploader.upload(photo)

            filename = result["secure_url"]
            photo_source = "cloudinary"
            cloudinary_public_id = result["public_id"]

        else:

            os.makedirs(
                app.config["UPLOAD_FOLDER"],
                exist_ok=True
            )

            filename = secure_filename(
                str(uuid.uuid4()) + "_" + photo.filename
            )

            photo.save(
                os.path.join(
                    app.config["UPLOAD_FOLDER"],
                    filename
                )
            )

            photo_source = "local"

    # ============================
    # Save Employee
    # ============================

    new_employee = Employee(
        employee_no=employee_no,
        first_name=first_name,
        last_name=last_name,
        email=email,
        department=department,
        position=position,
        status=status,
        photo=filename,
        photo_source=photo_source,
        cloudinary_public_id=cloudinary_public_id
    )

    db.session.add(new_employee)

    add_audit_log(
        f"Added Employee {employee_no}"
    )

    db.session.commit()

    flash(
        "Employee Added Successfully!",
        "success"
    )

    return redirect("/")

@app.route('/delete/<int:id>')
def delete_employee(id):
    if 'username' not in session:
        return redirect('/login')

    employee = Employee.query.get_or_404(id)
    if employee.photo_source == "local":
       if employee.photo:
           
           photo_path = os.path.join(
               app.config["UPLOAD_FOLDER"],
               employee.photo
           )

           if os.path.exists(photo_path):
               os.remove(photo_path)
    elif employee.photo_source == "cloudinary":
        if employee.cloudinary_public_id:
            cloudinary.uploader.destroy(
                employee.cloudinary_public_id
            )  

    db.session.delete(employee)
    add_audit_log(
        f"Deleted Employee {employee.employee_no}"
    )
    db.session.commit()
    flash(
        'Employee Deleted Successfully',
        'success'

    )
    return redirect('/')

@app.route('/edit/<int:id>', methods=['GET', 'POST'])

def edit_employee(id):
    if 'username' not in session:
        return redirect('/login')
    
    employee= Employee.query.get_or_404(id)
    if request.method == 'POST':

        employee.employee_no = request.form['employee_no']
        employee.first_name = request.form['first_name']
        employee.last_name = request.form['last_name']
        employee.email = request.form['email']
        employee.department = request.form['department']
        employee.position = request.form['position']
        employee.status = request.form['status']
        
        photo = request.files['photo']
        if photo and photo.filename:

            is_cloud = os.getenv("IS_CLOUD", "").strip().lower()

            if is_cloud == "true":

                if employee.cloudinary_public_id:
                    cloudinary.uploader.destroy(
                        employee.cloudinary_public_id
                    )

                result = cloudinary.uploader.upload(photo)

                employee.photo = result["secure_url"]
                employee.photo_source = "cloudinary"
                employee.cloudinary_public_id = result["public_id"]

            else:

                if employee.photo_source == "local" and employee.photo:

                    old_photo_path = os.path.join(
                    app.config["UPLOAD_FOLDER"],
                    employee.photo
                    )

                    if os.path.exists(old_photo_path):
                        os.remove(old_photo_path)

                os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

                filename = secure_filename(
                    str(uuid.uuid4()) + "_" + photo.filename
                    )

                photo.save(
                    os.path.join(
                    app.config["UPLOAD_FOLDER"],
                    filename
                )
            )

                employee.photo = filename
                employee.photo_source = "local"
                employee.cloudinary_public_id = None
        
                
        add_audit_log(
            f"Updated Employee {employee.employee_no}"
        )


        db.session.commit()
        print('saved')

        return redirect('/')
    
    return render_template(
        'edit.html',
        employee=employee
    )

@app.route('/create-admin')
def create_admin():

    #return 'Disabled'

    hashed_password = generate_password_hash('admin123')

    admin = User(
        username='admin',
        password=hashed_password
    )
    db.session.add(admin)
    db.session.commit()

    return 'Admin Created'

@app.route('/employee/<int:id>')
def employee_details(id):

    if 'username' not in session:
        return redirect('/login')
    
    employee = Employee.query.get_or_404(id)

    return render_template(
        'employee_details.html',
        employee=employee
    )

@app.route('/login', methods=['GET','POST'])
def login():
    if 'username' in session:
        return redirect('/')

    if request.method == "POST":

        username = request.form['username']
        password = request.form['password']

        user = User.query.filter_by(
            username=username
        ).first()

        if user and check_password_hash(
            user.password,
            password
        ):
            session['username'] = user.username
            log = AuditLog(
                username=user.username,
                action="Logged In"
            )
            db.session.add(log)
            db.session.commit()
            return redirect('/')
        flash(
            'Invalid Username or Password',
            'danger'
        )
    return render_template('login.html')

@app.route('/logout')
def logout():

    if 'username' in session:
        log = AuditLog(
            username=session['username'],
            action='Logged Out'
        )
        db.session.add(log)
        db.session.commit()
        
    session.pop('username',None)
    return redirect('/login')


@app.route('/export')
def export_employees():
    if 'username' not in session:
        return redirect('/login')
    
    employees = Employee.query.all()

    wb = Workbook()
    ws= wb.active

    ws.title = "Employees"
    ws['A1'] = "Employee Management System"
    ws['A1'].font = Font(
        bold=True,
        size=14
    )
    ws['A3'] = f"Generated By: {session['username']}"
    ws['A4'] = (
        "Generated Date: " + 
        datetime.now().strftime(
            "%m/%d/%Y %I:%M %p"
        )
    )
    headers = [
        'Employee No',
        'First Name',
        'Last Name',
        'Email',
        'Department',
        'Position',
        'Status'
    ]
    for col_num, header in enumerate(
        headers,
        start=1
    ):
        cell = ws.cell(
            row=6,
            column=col_num
        )
        cell.value = header
        cell.font = Font(
            bold=True
        )

    row = 7 

    for employee in employees:
    
        ws.cell(
            row=row,
            column=1,
            value=employee.employee_no
        )
        ws.cell(
            row=row,
            column=2,
            value=employee.first_name
        )
        ws.cell(
            row=row,
            column=3,
            value=employee.last_name
        )
        ws.cell(
            row=row,
            column=4,
            value=employee.email
        )
        ws.cell(
            row=row,
            column=5,
            value=employee.department
        )
        ws.cell(
            row=row,
            column=6,
            value=employee.position
        )
        ws.cell(
            row=row,
            column=7,
            value=employee.status
        )
        row += 1

    for column in  ws.columns:
        max_length = 0 
        
        column_letter = (
            column[0].column_letter
        )

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (
            max_length + 2
        )
        ws.column_dimensions[
            column_letter
        ].width = adjusted_width


    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='employees.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    )

if __name__ == '__main__':

    #with app.app_context():
       #db.create_all()
      app.run(debug=True)